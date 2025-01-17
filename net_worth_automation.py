import os
import shutil
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook, Workbook
from datetime import datetime

# Paths
DOWNLOADS_FOLDER = str(os.path.join(Path.home(), "Downloads"))
PROCESSED_FOLDER = str(os.path.join(os.getcwd(), "transactions"))
NEW_TRANSACTIONS_FOLDER = str(os.path.join(os.getcwd(), "new_transactions"))
EXCEL_WORKBOOK_PATH = str(os.path.join(os.getcwd(), "Monthly Net Worth Tracker.xlsx"))
SETUP_CSV_PATH = str(os.path.join(os.getcwd(), "setup.csv"))

# File name patterns to identify relevant CSVs
FILE_PATTERNS = ["transactions", "bank", "credit", "report"]

# Create folders if they don't exist
os.makedirs(PROCESSED_FOLDER, exist_ok=True)
os.makedirs(NEW_TRANSACTIONS_FOLDER, exist_ok=True)

def move_csv_files(start_dir:str, end_dir:str) -> None:
    """
    Move relevant CSV files from the given directory to the end directory.

    Args:
        start_dir: the directory to move the csv files from.
        end_dir: the directory to move the csv file to.
    Returns: 
        None
    """
    for file_name in os.listdir(start_dir):
        if any(pattern in file_name.lower() for pattern in FILE_PATTERNS) and file_name.endswith(".csv"):
            shutil.move(os.path.join(start_dir, file_name), os.path.join(end_dir, file_name))


def read_transactions(dir:str):
    """
    Read all transactions from CSVs in the given folder.
    
    Args:
        dir: the directory to read CSV transactional data from.
    Returns:
        None"""
    
    # Get the setup information
    setup = pd.read_csv(SETUP_CSV_PATH)

    transactions = pd.DataFrame()

    for index, row in setup.iterrows():
        # get the filename and file information
        file_name = row["file name"]
        debit_column = row["debit"]
        credit_column = row["credit"]
        check_type = row["check transaction type"]
        date = row["date"]
        description = row["description"]
        transaction_type = row["type"]
        header = row["header"]
        

        if file_name.endswith(".csv"):
            file_path = os.path.join(dir, file_name)
            
            if not header:
                df = pd.read_csv(file_path, header=None)

            else:
                df = pd.read_csv(file_path)
                
            if check_type:
                # Go through every transaction in the df and make the payments negative and income positive
                df.loc[df["type"] == "Debit", int(debit_column) - 1] *= -1

            # Combine credit and debit columns
            df["combined"] = df[debit_column - 1] - df[credit_column - 1]

            # Rearange columns to be in the correct order
            df = df.iloc[:, [date - 1, df.columns.get_loc("combined"), description - 1, transaction_type - 1]]

            # Rename columns
            df.columns = ["Date", "Amount", "Description", "Type"]

            # Add this to the dataframe of transactions
            transactions = pd.concat([transactions, df], ignore_index=True)

    return transactions


def update_net_worth_tracker(transactions):
    """Update the Excel workbook with the latest data."""
    # Load the workbook and get the sheet names
    try:
        wb = load_workbook(EXCEL_WORKBOOK_PATH)
    except FileNotFoundError as e:
        print(e)
        # TODO Create a new workbook
        # wb = Workbook()

    sheet_names = wb.sheetnames
    last_month_sheet = sheet_names[-1]
    print(sheet_names)
    print(last_month_sheet)

    # current_month_sheet = datetime.now().strftime("%B %Y")
    
    # # Copy the last month's sheet to a new sheet for the current month
    # wb.copy_worksheet(wb[last_month_sheet]).title = current_month_sheet
    
    # # Access the new sheet and update values
    # ws = wb[current_month_sheet]
    
    # # Assume specific cells to update; adjust according to your workbook structure
    # # Example: update total assets, liabilities, and net worth
    # current_assets = transactions['Amount'].sum()  # Adjust logic based on your file's structure
    # ws['B2'] = current_assets  # Example cell for total assets
    
    # # Save the workbook
    # wb.save(EXCEL_WORKBOOK_PATH)


def generate_reports():
    """Generate monthly and yearly reports."""
    # Load the workbook
    wb = load_workbook(EXCEL_WORKBOOK_PATH, data_only=True)
    
    # Example: Generate yearly summary as a DataFrame
    yearly_data = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        total_assets = ws['B2'].value  # Adjust cell references based on workbook structure
        net_worth = ws['B3'].value    # Example: Net worth in cell B3
        yearly_data.append({"Month": sheet_name, "Total Assets": total_assets, "Net Worth": net_worth})
    
    report_df = pd.DataFrame(yearly_data)
    
    # Save the report as a new sheet in the workbook
    with pd.ExcelWriter(EXCEL_WORKBOOK_PATH, engine='openpyxl', mode='a') as writer:
        report_df.to_excel(writer, sheet_name="Yearly Report", index=False)


def main():
    """Main function to automate the net worth tracker."""
    # move_csv_files(DOWNLOADS_FOLDER, NEW_TRANSACTIONS_FOLDER)
    transactions = read_transactions(NEW_TRANSACTIONS_FOLDER)
    update_net_worth_tracker(transactions)
    # generate_reports()
    # print("Net worth tracker updated successfully!")

if __name__ == "__main__":
    main()
